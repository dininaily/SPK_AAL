from flask import Flask, request, send_file, jsonify, session, redirect, url_for, render_template
from flask_cors import CORS
import io
import pandas as pd
import numpy as np
import warnings
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.cell.cell import MergedCell

warnings.filterwarnings('ignore')

app = Flask(__name__)
CORS(app)
app.secret_key = os.environ.get('SECRET_KEY', 'ganti-dengan-kunci-rahasia-acak')

PASSWORD_SPK = os.environ.get('SPK_PASSWORD', 'subditmindik2026')


COL_NO       = 0
COL_NAMA     = 1
COL_NO_AK    = 2
COL_PSI_I    = 3
COL_PSI_II   = 4
COL_PSI_III  = 5
COL_PSI_IV   = 6
COL_PSI_V    = 7
COL_STAKES   = 9
COL_KESWA    = 10
COL_JAZ_AB   = 12
COL_JAZ_RNG  = 13
COL_LAT_KAP  = 16
COL_LAT_SEA  = 17
COL_LAT_PRA  = 18
COL_AKD_ING  = 21
COL_AKD_MAT  = 22
COL_AKD_FIS  = 23
COL_IPK      = 26
COL_IPK_BAKU = 26
COL_PIL_I    = 28
COL_PIL_II   = 29
COL_PIL_III  = 30
COL_PIL_IV   = 31
COL_PIL_V    = 32

KORPS = ['P', 'T', 'E', 'S', 'M']
KORPS_NAMA = {
    'P': 'Pelaut',
    'T': 'Teknik',
    'E': 'Elektro',
    'S': 'Suplai',
    'M': 'Marinir',
}


def safe_str(val):
    s = str(val).strip()
    return '' if s.lower() == 'nan' else s

def konversi_c1(psi_i, psi_ii, psi_iii, psi_iv, psi_v, korps):
    skor_map = [5, 4, 3, 1, 1]
    for urutan, val in enumerate([psi_i, psi_ii, psi_iii, psi_iv, psi_v]):
        if safe_str(val) == korps:
            return skor_map[urutan]
    return 1

def konversi_stakes(val):
    mapping = {'I': 5, 'II': 4, 'III': 3, 'IV': 1}
    return mapping.get(str(val).strip(), 1)

def konversi_keswa(val):
    mapping = {'J1': 5, 'J2': 4, 'J3': 3}
    return mapping.get(str(val).strip(), 3)

def konversi_c2(stakes_raw, keswa_raw):
    return (konversi_stakes(stakes_raw) + konversi_keswa(keswa_raw)) / 2

def konversi_c3(baterai_ab, renang):
    try:
        rata = (float(baterai_ab) + float(renang)) / 2
    except:
        return 1
    if rata >= 90:   return 5
    elif rata >= 80: return 4
    elif rata >= 76: return 3
    elif rata >= 70: return 2
    else:            return 1

def konversi_c4(kapal, sea, pra):
    try:
        rata = (float(kapal) + float(sea) + float(pra)) / 3
    except:
        return 1
    if rata >= 85:   return 5
    elif rata >= 75: return 4
    elif rata >= 65: return 3
    elif rata >= 55: return 2
    else:            return 1

def konversi_c5(ing, mat, fis):
    try:
        rata = (float(ing) + float(mat) + float(fis)) / 3
    except:
        return 1
    if rata > 89.9:  return 5
    elif rata >= 80: return 4
    elif rata >= 75: return 3
    elif rata >= 70: return 2
    else:            return 1

def konversi_c6(ipk):
    try:
        ipk = float(ipk)
    except:
        return 1
    if ipk >= 3.60:   return 5
    elif ipk >= 3.00: return 4
    elif ipk >= 2.26: return 3
    elif ipk >= 1.96: return 2
    else:             return 1

def konversi_c7(pil_i, pil_ii, pil_iii, pil_iv, pil_v, korps):
    pilihan = [safe_str(p) for p in [pil_i, pil_ii, pil_iii, pil_iv, pil_v]]
    skor_map = {0: 5, 1: 4, 2: 3, 3: 2, 4: 1}
    for i, p in enumerate(pilihan):
        if p == korps:
            return skor_map[i]
    return 1


def baca_data(file_bytes):
    df_raw = pd.read_excel(io.BytesIO(file_bytes), sheet_name=0, header=None)
    data = df_raw.iloc[6:].reset_index(drop=True)
    data = data[pd.to_numeric(data[COL_NO], errors='coerce').notna()].reset_index(drop=True)
    return data

def hitung_entropy_bobot(matriks):
    n = matriks.shape[0]
    col_sum = matriks.sum(axis=0)
    col_sum[col_sum == 0] = 1e-10
    P = matriks / col_sum
    with np.errstate(divide='ignore', invalid='ignore'):
        ln_P = np.where(P > 0, np.log(P), 0)
    E = -(1 / np.log(n)) * (P * ln_P).sum(axis=0)
    E = np.clip(E, 0, 1)
    d = 1 - E
    total_d = d.sum()
    if total_d == 0:
        bobot = np.ones(7) / 7
    else:
        bobot = d / total_d
    return bobot

def normalisasi_saw(matriks):
    col_max = matriks.max(axis=0)
    col_max[col_max == 0] = 1e-10
    return matriks / col_max

def proses(data_bytes):
    data = baca_data(data_bytes)
    n = len(data)

    skor_c2 = np.array([konversi_c2(data.iloc[i][COL_STAKES], data.iloc[i][COL_KESWA]) for i in range(n)])
    skor_c3 = np.array([konversi_c3(data.iloc[i][COL_JAZ_AB], data.iloc[i][COL_JAZ_RNG]) for i in range(n)])
    skor_c4 = np.array([konversi_c4(data.iloc[i][COL_LAT_KAP], data.iloc[i][COL_LAT_SEA], data.iloc[i][COL_LAT_PRA]) for i in range(n)])
    skor_c5 = np.array([konversi_c5(data.iloc[i][COL_AKD_ING], data.iloc[i][COL_AKD_MAT], data.iloc[i][COL_AKD_FIS]) for i in range(n)])
    skor_c6 = np.array([konversi_c6(data.iloc[i][COL_IPK]) for i in range(n)])

    hasil = []
    bobot_per_korps = {}

    for korps in KORPS:
        skor_c1 = np.array([konversi_c1(
            data.iloc[i][COL_PSI_I], data.iloc[i][COL_PSI_II],
            data.iloc[i][COL_PSI_III], data.iloc[i][COL_PSI_IV],
            data.iloc[i][COL_PSI_V], korps) for i in range(n)])
        skor_c7 = np.array([konversi_c7(
            data.iloc[i][COL_PIL_I], data.iloc[i][COL_PIL_II],
            data.iloc[i][COL_PIL_III], data.iloc[i][COL_PIL_IV],
            data.iloc[i][COL_PIL_V], korps) for i in range(n)])

        matriks = np.column_stack([skor_c1, skor_c2, skor_c3, skor_c4, skor_c5, skor_c6, skor_c7])
        bobot = hitung_entropy_bobot(matriks)
        bobot_per_korps[korps] = bobot

        R = normalisasi_saw(matriks)
        V = (R * bobot).sum(axis=1)

        for i in range(n):
            if len(hasil) <= i:
                hasil.append({
                    'No':           int(data.iloc[i][COL_NO]),
                    'Nama':         data.iloc[i][COL_NAMA],
                    'No_AK':        safe_str(data.iloc[i][COL_NO_AK]),
                    'PSI_I':        safe_str(data.iloc[i][COL_PSI_I]),
                    'PSI_II':       safe_str(data.iloc[i][COL_PSI_II]),
                    'PSI_III':      safe_str(data.iloc[i][COL_PSI_III]),
                    'PSI_IV':       safe_str(data.iloc[i][COL_PSI_IV]),
                    'PSI_V':        safe_str(data.iloc[i][COL_PSI_V]),
                    'STAKES_RAW':   safe_str(data.iloc[i][COL_STAKES]),
                    'STAKES_KONV':  konversi_stakes(data.iloc[i][COL_STAKES]),
                    'KESWA_RAW':    safe_str(data.iloc[i][COL_KESWA]),
                    'KESWA_KONV':   konversi_keswa(data.iloc[i][COL_KESWA]),
                    'C2':           skor_c2[i],
                    'JAZ_AB':       data.iloc[i][COL_JAZ_AB],
                    'JAZ_RNG':      data.iloc[i][COL_JAZ_RNG],
                    'C3':           int(skor_c3[i]),
                    'LAT_KAP':      data.iloc[i][COL_LAT_KAP],
                    'LAT_SEA':      data.iloc[i][COL_LAT_SEA],
                    'LAT_PRA':      data.iloc[i][COL_LAT_PRA],
                    'C4':           int(skor_c4[i]),
                    'AKD_ING':      data.iloc[i][COL_AKD_ING],
                    'AKD_MAT':      data.iloc[i][COL_AKD_MAT],
                    'AKD_FIS':      data.iloc[i][COL_AKD_FIS],
                    'C5':           int(skor_c5[i]),
                    'IPK_RAW':      data.iloc[i][COL_IPK],
                    'IPK_BAKU':     data.iloc[i][COL_IPK_BAKU],
                    'C6':           int(skor_c6[i]),
                    'PIL_I':        safe_str(data.iloc[i][COL_PIL_I]),
                    'PIL_II':       safe_str(data.iloc[i][COL_PIL_II]),
                    'PIL_III':      safe_str(data.iloc[i][COL_PIL_III]),
                    'PIL_IV':       safe_str(data.iloc[i][COL_PIL_IV]),
                    'PIL_V':        safe_str(data.iloc[i][COL_PIL_V]),
                })

            hasil[i][f'C1_{KORPS_NAMA[korps]}'] = int(skor_c1[i])
            hasil[i][f'C7_{KORPS_NAMA[korps]}'] = int(skor_c7[i])
            hasil[i][f'V_{KORPS_NAMA[korps]}']  = round(V[i], 4)

    for row in hasil:
        v_vals = {k: row[f'V_{KORPS_NAMA[k]}'] for k in KORPS}
        best_korps = max(v_vals, key=v_vals.get)
        row['Rekomendasi'] = KORPS_NAMA[best_korps]

    return hasil, bobot_per_korps


def buat_output_excel(hasil, bobot_per_korps, template_bytes, angkatan):
    wb = load_workbook(io.BytesIO(template_bytes))
    ws = wb.active

    ws['A3'] = (
        f"PENILAIAN TERHADAP 7 KOMPONEN DALAM RANGKA "
        f"PENENTUAN KORPS TARUNA AAL ANGKATAN KE-{angkatan}"
    )

    BARIS_MULAI = 8

    def tulis(baris, col, val, center=True):
        cell = ws.cell(row=baris, column=col)
        if isinstance(cell, MergedCell):
            return
        cell.value = val
        cell.font = Font(size=9, name='Arial')
        cell.alignment = Alignment(
            horizontal='center' if center else 'left',
            vertical='center'
        )

    for i, row_data in enumerate(hasil):
        baris = BARIS_MULAI + i
        t = lambda col, val, center=True: tulis(baris, col, val, center)

        t(1, row_data['No'])
        t(2, row_data['Nama'], False)
        t(3, row_data['No_AK'])

        t(4, row_data['PSI_I'])
        t(5, row_data['PSI_II'])
        t(6, row_data['PSI_III'])
        t(7, row_data['PSI_IV'])
        t(8, row_data['PSI_V'])

        t(9,  row_data['C1_Pelaut'])
        t(10, row_data['C1_Teknik'])
        t(11, row_data['C1_Elektro'])
        t(12, row_data['C1_Suplai'])
        t(13, row_data['C1_Marinir'])

        t(15, row_data['STAKES_RAW'])
        t(16, row_data['STAKES_KONV'])
        t(17, row_data['KESWA_RAW'])
        t(18, row_data['KESWA_KONV'])
        t(19, row_data['C2'])

        t(21, row_data['JAZ_AB'])
        t(22, row_data['JAZ_RNG'])
        try:
            rata_jaz = round((float(row_data['JAZ_AB']) + float(row_data['JAZ_RNG'])) / 2, 2)
        except:
            rata_jaz = ''
        t(23, rata_jaz)
        t(24, row_data['C3'])

        t(26, row_data['LAT_KAP'])
        t(27, row_data['LAT_SEA'])
        t(28, row_data['LAT_PRA'])
        try:
            rata_lat = round((float(row_data['LAT_KAP']) + float(row_data['LAT_SEA']) + float(row_data['LAT_PRA'])) / 3, 2)
        except:
            rata_lat = ''
        t(29, rata_lat)
        t(30, row_data['C4'])

        t(32, row_data['AKD_ING'])
        t(33, row_data['AKD_MAT'])
        t(34, row_data['AKD_FIS'])
        try:
            rata_akd = round((float(row_data['AKD_ING']) + float(row_data['AKD_MAT']) + float(row_data['AKD_FIS'])) / 3, 2)
        except:
            rata_akd = ''
        t(35, rata_akd)
        t(36, row_data['C5'])

        t(38, row_data['IPK_RAW'])
        t(39, row_data['C6'])

        t(41, row_data['PIL_I'])
        t(42, row_data['PIL_II'])
        t(43, row_data['PIL_III'])
        t(44, row_data['PIL_IV'])
        t(45, row_data['PIL_V'])

        t(46, row_data['C7_Pelaut'])
        t(47, row_data['C7_Teknik'])
        t(48, row_data['C7_Elektro'])
        t(49, row_data['C7_Suplai'])
        t(50, row_data['C7_Marinir'])

        t(52, row_data['V_Pelaut'])
        t(53, row_data['V_Teknik'])
        t(54, row_data['V_Elektro'])
        t(55, row_data['V_Suplai'])
        t(56, row_data['V_Marinir'])

        v_vals = [(KORPS_NAMA[k], row_data[f'V_{KORPS_NAMA[k]}']) for k in KORPS]
        v_sorted = sorted(v_vals, key=lambda x: x[1], reverse=True)
        saran_cols = [57, 59, 61, 63, 65]
        for idx, (nama, v_val) in enumerate(v_sorted):
            col_val = saran_cols[idx]
            t(col_val,     round(v_val, 4))
            t(col_val + 1, nama[0])

        ws.row_dimensions[baris].height = 16

    if "Bobot Entropy" in wb.sheetnames:
        ws2 = wb["Bobot Entropy"]
    else:
        ws2 = wb.create_sheet("Bobot Entropy")

    ws2['A1'] = 'BOBOT ENTROPY PER KRITERIA PER KORPS'
    ws2['A1'].font = Font(bold=True, size=12, name='Arial')

    for c, h in enumerate(['Kriteria', 'Pelaut', 'Teknik', 'Elektro', 'Suplai', 'Marinir'], start=1):
        cell = ws2.cell(row=2, column=c, value=h)
        cell.font = Font(bold=True, color='FFFFFF', size=10, name='Arial')
        cell.fill = PatternFill('solid', start_color='1A3A5C')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    kriteria_nama = [
        'C1 - Psikologi', 'C2 - Kesehatan', 'C3 - Jasmani',
        'C4 - Latihan', 'C5 - Akademik', 'C6 - Kepribadian',
        'C7 - Pilihan Korps'
    ]
    for r, krit in enumerate(kriteria_nama, start=3):
        ws2.cell(row=r, column=1, value=krit).font = Font(size=9, bold=True, name='Arial')
        for c, korps in enumerate(KORPS, start=2):
            cell = ws2.cell(row=r, column=c, value=round(float(bobot_per_korps[korps][r - 3]), 4))
            cell.font = Font(size=9, name='Arial')
            cell.alignment = Alignment(horizontal='center')

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


@app.route('/proses', methods=['POST'])
def route_proses():
    if 'template' not in request.files or 'data' not in request.files:
        return jsonify({'error': 'File template dan data wajib diupload'}), 400

    template_bytes = request.files['template'].read()
    data_bytes     = request.files['data'].read()
    angkatan       = request.form.get('angkatan', '').strip()

    if not angkatan:
        angkatan = 'X'

    try:
        hasil, bobot_per_korps = proses(data_bytes)
    except Exception as e:
        return jsonify({'error': f'Gagal memproses data: {str(e)}'}), 500

    try:
        output = buat_output_excel(hasil, bobot_per_korps, template_bytes, angkatan)
    except Exception as e:
        return jsonify({'error': f'Gagal menulis ke template: {str(e)}'}), 500

    filename = f'Hasil_Jurkorps_{angkatan}.xlsx'
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

from flask import render_template
@app.route('/')
def index():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    return render_template('SPK_AAL_JURKORPS.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        password = request.form.get('password', '')
        if password == PASSWORD_SPK:
            session['logged_in'] = True
            return redirect(url_for('index'))
        return render_template('login.html', error='Password salah')
    return render_template('login.html', error=None)

@app.route('/logout')
def logout():
    session.pop('logged_in', None)
    return redirect(url_for('login'))

@app.route('/ping', methods=['GET'])
def ping():
    return jsonify({'status': 'ok'})

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=False, host='0.0.0.0', port=port)
